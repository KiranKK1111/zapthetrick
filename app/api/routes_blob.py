"""Serve + preview any stored blob (uploads AND generated artifacts).

Everything lives in the Postgres blob store, so the same two endpoints back the
right-docked preview panel for every supported file type:

  GET /api/blob?path=...           -> raw bytes (images shown directly; any
                                      download).
  GET /api/blob/preview?path=...   -> {kind, ...} telling the FE how to render:
        image | pdf (rasterized pages) | text | download
      Almost everything is previewable as text — plain text/code/data inline,
      and PowerPoint / OpenDocument / archives via extracted text. Only Word &
      Excel (.docx/.doc/.xlsx/.xls) and unreadable binaries are download-only.
"""
from __future__ import annotations

import base64
import mimetypes

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse, Response

from storage.blobs.factory import get_blobs

router = APIRouter(tags=["blob"])

_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
# Download-only — Word + legacy Excel. .xlsx now previews in the grid (below);
# .xls (old binary) openpyxl can't read, so it stays download-only.
_DOWNLOAD_EXT = {".docx", ".doc", ".xls"}
# Excel we CAN parse → a spreadsheet-grid preview (rich-viewers #3).
_SPREADSHEET_EXT = {".xlsx"}
# Formats that aren't plain text but whose READABLE TEXT we can extract and show
# inline (PowerPoint, OpenDocument, and archives → their members' text).
_EXTRACT_EXT = {
    ".pptx", ".ppt", ".odt", ".ods", ".odp",
    ".zip", ".7z", ".rar", ".tar", ".tgz", ".tbz2", ".txz", ".tzst",
    ".gz", ".bz2", ".xz", ".zst", ".lz4", ".br",
}
# Text + every common source/config/data type -> inline text preview.
_TEXT_EXT = {
    ".txt": "text", ".log": "text", ".md": "markdown", ".markdown": "markdown",
    ".rst": "text",
    ".json": "json", ".json5": "json", ".jsonc": "json",
    ".csv": "csv", ".tsv": "csv", ".adoc": "text",
    ".yaml": "code", ".yml": "code", ".toml": "code", ".ini": "code",
    ".cfg": "code", ".conf": "code", ".properties": "code", ".env": "code",
    ".xml": "code", ".xsd": "code", ".xsl": "code", ".xslt": "code",
    ".html": "code", ".htm": "code", ".xhtml": "code",
    ".css": "code", ".scss": "code", ".sass": "code", ".less": "code", ".styl": "code", ".svg": "code",
    ".tf": "code", ".tfvars": "code", ".hcl": "code", ".bicep": "code",
    ".cmake": "code", ".mk": "code", ".gradle": "code", ".sbt": "code",
    ".lock": "code", ".sum": "code", ".mod": "code",
    ".js": "code", ".jsx": "code", ".ts": "code", ".tsx": "code", ".mjs": "code", ".cjs": "code",
    ".py": "code", ".pyi": "code", ".pyx": "code", ".rb": "code", ".php": "code",
    ".pl": "code", ".pm": "code", ".lua": "code",
    ".java": "code", ".kt": "code", ".kts": "code", ".scala": "code", ".groovy": "code",
    ".c": "code", ".h": "code", ".cpp": "code", ".cc": "code", ".cxx": "code",
    ".hpp": "code", ".hh": "code", ".cs": "code", ".go": "code", ".rs": "code", ".swift": "code",
    ".dart": "code", ".m": "code", ".mm": "code", ".r": "code", ".jl": "code",
    ".asm": "code", ".s": "code", ".vb": "code", ".fs": "code", ".fsx": "code", ".fsi": "code",
    ".ml": "code", ".mli": "code", ".hs": "code", ".lhs": "code", ".erl": "code", ".hrl": "code",
    ".ex": "code", ".exs": "code", ".clj": "code", ".cljs": "code", ".cljc": "code", ".edn": "code",
    ".lisp": "code", ".el": "code", ".scm": "code", ".rkt": "code", ".nim": "code", ".zig": "code",
    ".d": "code", ".pas": "code", ".pp": "code", ".f90": "code", ".f95": "code",
    ".cob": "code", ".cbl": "code", ".tcl": "code", ".vhd": "code", ".vhdl": "code",
    ".v": "code", ".sv": "code", ".svh": "code", ".coffee": "code", ".sol": "code", ".ino": "code",
    ".sh": "code", ".bash": "code", ".zsh": "code", ".fish": "code",
    ".ps1": "code", ".psm1": "code", ".bat": "code", ".cmd": "code",
    ".sql": "code", ".graphql": "code", ".gql": "code", ".prisma": "code",
    ".proto": "code", ".thrift": "code", ".vue": "code", ".svelte": "code", ".astro": "code",
    ".patch": "code", ".diff": "code",
    ".gitignore": "code", ".editorconfig": "code", ".mako": "code", ".ipynb": "json",
}
_MAX_TEXT_CHARS = 800_000
_PDF_MAX_PAGES = 40

# Only blobs under these known prefixes may be served — stops path-traversal /
# blind enumeration of arbitrary store keys (the filesystem-fallback store would
# otherwise be reachable via `..`). Every writer in the app uses one of these.
_ALLOWED_PREFIXES = (
    "chat_images/", "documents/", "resumes/", "solve/",
)


def _path_ok(path: str) -> bool:
    p = (path or "").replace("\\", "/")
    if not p or ".." in p or p.startswith("/") or ":" in p:
        return False
    return any(p.startswith(prefix) for prefix in _ALLOWED_PREFIXES)


def _ext(name: str) -> str:
    i = name.rfind(".")
    return name[i:].lower() if i >= 0 else ""


def _looks_texty(data: bytes) -> bool:
    """Heuristic: does this look like a text/source file? (no NUL bytes, mostly
    printable). Lets unknown source extensions still preview as text."""
    sample = data[:8192]
    if not sample:
        return True
    if b"\x00" in sample:
        return False
    good = sum(1 for b in sample if b in (9, 10, 13) or 32 <= b <= 126 or b >= 128)
    return good / len(sample) > 0.85


@router.get("/blob")
async def serve_blob(path: str = Query(...)):
    """Stream a blob's raw bytes with a best-effort content type."""
    if not _path_ok(path):
        return Response(status_code=404)
    try:
        data = await get_blobs().get(path)
    except FileNotFoundError:
        return Response(status_code=404)
    except Exception:  # noqa: BLE001 — blob store not ready
        return Response(status_code=503)
    ctype = mimetypes.guess_type(path)[0] or "application/octet-stream"
    return Response(content=data, media_type=ctype,
                    headers={"Cache-Control": "no-store"})


def _rasterize_pdf(data: bytes) -> list[str]:
    """PDF -> list of base64 PNG pages (PyMuPDF)."""
    import fitz  # PyMuPDF

    out: list[str] = []
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        mat = fitz.Matrix(1.6, 1.6)
        for i, page in enumerate(doc):
            if i >= _PDF_MAX_PAGES:
                break
            pix = page.get_pixmap(matrix=mat, alpha=False)
            out.append(base64.b64encode(pix.tobytes("png")).decode())
    finally:
        doc.close()
    return out


def _xlsx_preview_tables(data: bytes, max_rows: int = 300, max_cols: int = 40) -> str:
    """Render an .xlsx as markdown pipe tables — one per sheet — for the FE's
    spreadsheet grid (which shows one tab per table). Cells are bounded and any
    ``|`` / newlines neutralised so they can't break the table columns."""
    import io as _io

    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    blocks: list[str] = []
    try:
        for ws in wb.worksheets:
            rows_md: list[str] = []
            for r, row in enumerate(ws.iter_rows(values_only=True)):
                if r >= max_rows:
                    break
                cells = [
                    ("" if c is None else str(c)).replace("|", "/").replace("\n", " ")
                    for c in row[:max_cols]
                ]
                if not any(c.strip() for c in cells):
                    continue
                rows_md.append("| " + " | ".join(cells) + " |")
            if rows_md:
                blocks.append("\n".join(rows_md))
    finally:
        wb.close()
    return "\n\n".join(blocks).strip()


def _route(ext: str) -> tuple[str, str | None]:
    """Pure routing decision for a file extension (single source of truth, so it
    can be unit-tested without I/O). Returns ``(kind, fmt)``:

      image | download | pdf | extract | text | sniff

    'text' carries the syntax format in ``fmt``; the others have ``fmt=None``.
    """
    if ext in _IMAGE_EXT:
        return ("image", None)
    if ext in _SPREADSHEET_EXT:        # .xlsx → grid preview
        return ("spreadsheet", None)
    if ext in _DOWNLOAD_EXT:           # Word / legacy Excel — download-only
        return ("download", None)
    if ext == ".pdf":
        return ("pdf", None)
    fmt = _TEXT_EXT.get(ext)
    if fmt is not None:                # known text/code/data → inline
        return ("text", fmt)
    if ext in _EXTRACT_EXT:            # PowerPoint / ODF / archives → extract
        return ("extract", None)
    return ("sniff", None)             # unknown → sniff bytes, text or download


@router.get("/blob/preview")
async def preview_blob(path: str = Query(...), name: str | None = Query(None)):
    """Classify a blob and return what the FE needs to render it in the panel."""
    if not _path_ok(path):
        return JSONResponse({"kind": "error", "detail": "not found"}, status_code=404)
    fname = name or path
    ext = _ext(fname)
    route, fmt = _route(ext)

    if route == "image":
        return {"kind": "image", "path": path, "name": fname}
    if route == "download":
        return {"kind": "download", "path": path, "name": fname}

    if route == "spreadsheet":
        try:
            data = await get_blobs().get(path)
        except FileNotFoundError:
            return JSONResponse({"kind": "error", "detail": "not found"}, status_code=404)
        try:
            tables = _xlsx_preview_tables(data)
        except Exception:  # noqa: BLE001 — any parse failure never breaks the panel
            tables = ""
        # Empty / unreadable workbook → fall back to a download card.
        if not tables:
            return {"kind": "download", "path": path, "name": fname}
        return {"kind": "spreadsheet", "text": tables, "name": fname}

    if route == "pdf":
        try:
            data = await get_blobs().get(path)
        except FileNotFoundError:
            return JSONResponse({"kind": "error", "detail": "not found"}, status_code=404)
        try:
            return {"kind": "pdf", "pages": _rasterize_pdf(data), "name": fname}
        except Exception as exc:  # noqa: BLE001
            return JSONResponse({"kind": "error", "detail": f"pdf render failed: {exc}"},
                                status_code=500)

    # Known text/code/data extension → ranged text preview (single scroll).
    # Ranged read: never pull more than we can show. utf-8 chars are ≥1 byte, so
    # _MAX_TEXT_CHARS*4 bytes always yields ≥ _MAX_TEXT_CHARS chars after decode
    # — a 100 MB blob is bounded to ~3.2 MB at the store, not in RAM.
    if route == "text":
        try:
            data = await get_blobs().get_prefix(path, _MAX_TEXT_CHARS * 4)
        except FileNotFoundError:
            return JSONResponse({"kind": "error", "detail": "not found"}, status_code=404)
        txt = data.decode("utf-8", errors="replace")[:_MAX_TEXT_CHARS]
        return {"kind": "text", "text": txt, "format": fmt, "name": fname}

    # PowerPoint / OpenDocument / archives: EXTRACT readable text and show it
    # inline (needs the full bytes to parse the container).
    if route == "extract":
        try:
            data = await get_blobs().get(path)
        except FileNotFoundError:
            return JSONResponse({"kind": "error", "detail": "not found"}, status_code=404)
        try:
            import asyncio

            from app.documents.parser import (  # noqa: PLC0415
                PasswordRequired,
                UnsupportedDocument,
                extract_document_text,
            )

            text = await asyncio.to_thread(extract_document_text, data, fname)
            if text and text.strip():
                return {"kind": "text", "text": text[:_MAX_TEXT_CHARS],
                        "format": "text", "name": fname}
        except (UnsupportedDocument, PasswordRequired):
            pass
        except Exception:  # noqa: BLE001 — extraction failed; fall back
            pass
        return {"kind": "download", "path": path, "name": fname}

    # Unknown extension: sniff a bounded prefix (don't pull a whole binary just
    # to check) — inline if it looks like text, else download.
    try:
        data = await get_blobs().get_prefix(path, _MAX_TEXT_CHARS * 4)
    except FileNotFoundError:
        return JSONResponse({"kind": "error", "detail": "not found"}, status_code=404)
    if _looks_texty(data):
        txt = data.decode("utf-8", errors="replace")[:_MAX_TEXT_CHARS]
        return {"kind": "text", "text": txt, "format": "text", "name": fname}
    return {"kind": "download", "path": path, "name": fname}
