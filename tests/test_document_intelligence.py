"""Phase 10 — document intelligence (#9).

Section numbering + TOC (PDF/DOCX), Excel numeric coercion + SUM totals row +
bar chart, and Mermaid labelling. Offline (openpyxl/docx/fpdf imported lazily).
"""
from __future__ import annotations

import io

from app.documents.generators import (
    _coerce_num,
    _numeric_columns,
    number_headings,
    parse_blocks,
    render_document,
)

MULTI_SECTION_MD = """\
# Introduction

Intro prose.

## Background

Some background.

## Goals

The goals.

# Implementation

Details here.
"""

TABLE_MD = """\
# Sales

| Region | Revenue | Units |
| --- | --- | --- |
| North | 1200 | 30 |
| South | 900 | 22 |
| East | 1500 | 40 |
"""

MERMAID_MD = """\
# Flow

```mermaid
graph TD; A-->B;
```
"""


# ── section numbering + TOC ─────────────────────────────────────────────────
def test_number_headings_assigns_hierarchy():
    blocks = parse_blocks(MULTI_SECTION_MD)
    numbered, toc = number_headings(blocks)
    heads = [b for b in numbered if b[0] == "h"]
    texts = [b[2] for b in heads]
    assert texts[0].startswith("1  Introduction")
    assert texts[1].startswith("1.1  Background")
    assert texts[2].startswith("1.2  Goals")
    assert texts[3].startswith("2  Implementation")
    assert len(toc) == 4
    assert toc[0] == (1, "1", "Introduction")
    assert toc[1] == (2, "1.1", "Background")


def test_number_headings_skipped_for_short_docs():
    blocks = parse_blocks("# Only One\n\nText.\n")
    numbered, toc = number_headings(blocks)
    assert toc == []
    assert numbered == blocks  # unchanged


def test_pdf_has_toc_and_numbered_sections():
    import pytest
    data, _, ext = render_document(MULTI_SECTION_MD, "pdf", title="Doc")
    assert ext == "pdf" and data[:5] == b"%PDF-"
    fitz = pytest.importorskip("fitz")
    doc = fitz.open(stream=data, filetype="pdf")
    text = "\n".join(p.get_text() for p in doc)
    doc.close()
    assert "Contents" in text
    assert "1.1" in text and "Implementation" in text


def test_docx_has_toc_and_numbered_sections():
    data, _, ext = render_document(MULTI_SECTION_MD, "docx")
    assert ext == "docx"
    from docx import Document
    doc = Document(io.BytesIO(data))
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "Contents" in text
    assert "1.1  Background" in text


# ── Excel numeric coercion + totals + chart ─────────────────────────────────
def test_coerce_num():
    assert _coerce_num("1200") == 1200
    assert _coerce_num("1,500") == 1500
    assert _coerce_num("3.14") == 3.14
    assert _coerce_num("8%") == 8
    assert _coerce_num("North") == "North"
    assert _coerce_num("") == ""


def test_numeric_columns_detection():
    rows = [["Region", "Revenue", "Units"],
            ["North", "1200", "30"], ["South", "900", "22"]]
    assert _numeric_columns(rows) == [1, 2]


def test_xlsx_has_numbers_totals_and_chart():
    data, _, ext = render_document(TABLE_MD, "xlsx", title="Sales")
    assert ext == "xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    # Header intact + bold (backward compat).
    assert [c.value for c in ws[1]] == ["Region", "Revenue", "Units"]
    assert ws[1][0].font.bold
    # Numeric cells coerced to numbers.
    assert ws["B2"].value == 1200 and isinstance(ws["B2"].value, int)
    # SUM totals row appended.
    last = ws.max_row
    assert ws.cell(row=last, column=1).value == "Total"
    assert str(ws.cell(row=last, column=2).value).startswith("=SUM(")
    # A chart was added.
    assert len(ws._charts) >= 1


def test_xlsx_non_numeric_table_no_chart():
    md = "| Name | City |\n| --- | --- |\n| A | NY |\n| B | LA |\n"
    data, _, _ = render_document(md, "xlsx")
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(data)).active
    assert len(ws._charts) == 0          # no numeric column → no chart
    assert ws.cell(row=ws.max_row, column=1).value != "Total"


# ── Mermaid labelling ───────────────────────────────────────────────────────
def test_parse_blocks_captures_code_language():
    blocks = parse_blocks(MERMAID_MD)
    code = [b for b in blocks if b[0] == "code"]
    assert code and code[0][2] == "mermaid"


def test_pdf_labels_mermaid():
    import pytest
    data, _, _ = render_document(MERMAID_MD, "pdf", title="Flow")
    fitz = pytest.importorskip("fitz")
    doc = fitz.open(stream=data, filetype="pdf")
    text = "\n".join(p.get_text() for p in doc)
    doc.close()
    assert "Mermaid" in text
